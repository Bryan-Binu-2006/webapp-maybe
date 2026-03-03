"""
utils/excel_helpers.py — Excel file processing
================================================
Handles reading an uploaded .xlsx file, extracting a single column
of participant names, and validating the data.

SECURITY & ROBUSTNESS NOTES:
- We use openpyxl (read_only=True) to stream rows without loading
  the entire workbook into memory.
- Maximum row count is enforced to prevent resource exhaustion.
- Each name is stripped of whitespace and validated for length.
- Empty names are silently skipped (they are not errors, just
  blank rows).
"""

import logging
from typing import List, Tuple

import openpyxl
from flask import current_app

logger = logging.getLogger(__name__)


def get_column_headers(file_path: str) -> List[str]:
    """
    Return the list of column headers (first row) from the first
    worksheet of the given .xlsx file.

    Used to let the user pick which column contains names.
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = [str(cell) if cell is not None else "" for cell in row]
    wb.close()
    return headers


def extract_names(
    file_path: str,
    column_index: int,
) -> Tuple[List[str], List[str]]:
    """
    Extract participant names from *column_index* (0-based) of
    the first worksheet.

    Returns:
        (valid_names, warnings)

    Validation rules:
    - Skip the header row (row 1).
    - Strip whitespace from each cell.
    - Skip empty cells.
    - Reject names longer than MAX_NAME_LENGTH.
    - Enforce MAX_BATCH_SIZE total rows.

    WHY these limits?
    - Extremely long names could break image rendering or be used
      for injection.
    - A huge row count could exhaust memory or disk space when
      generating certificate images.
    """
    max_rows = current_app.config["MAX_BATCH_SIZE"]
    max_name_len = current_app.config["MAX_NAME_LENGTH"]

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    names: List[str] = []
    warnings: List[str] = []
    row_num = 0

    for row in ws.iter_rows(min_row=2, values_only=True):  # skip header
        row_num += 1

        if row_num > max_rows:
            warnings.append(
                f"Stopped after {max_rows} rows (batch size limit)."
            )
            break

        # Safely get the cell value at the chosen column
        if column_index >= len(row):
            continue
        cell_value = row[column_index]
        if cell_value is None:
            continue

        name = str(cell_value).strip()
        if not name:
            continue

        if len(name) > max_name_len:
            warnings.append(
                f"Row {row_num + 1}: name too long ({len(name)} chars), skipped."
            )
            continue

        names.append(name)

    wb.close()

    logger.info("Extracted %d names from Excel file.", len(names))
    return names, warnings
